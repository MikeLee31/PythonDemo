import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import *

'''
打开excel
    参数为
返回一个workbook
'''


def open_excel(filepath):
    if os.path.exists(filepath):
        return load_workbook(filepath)
    else:
        Workbook().save(filepath)
        return load_workbook(filepath)


def open_sheet(workbook, sheetname):
    sheet_list = workbook.sheetnames
    if sheet_list.count(sheetname) == 1:
        return workbook[sheetname]
    else:
        return workbook.create_sheet(sheetname)


def get_value(worksheet):
    values =[]
    for row in ws.values:
        lis = []
        lis.clear()
        for value in row:
            lis.append(value)

        values.append(lis)
    return values


'''
# 访问A列4行的单元格，不存在则创建
c = ws['A4']
# 还有Worksheet.cell()方法，赋值（4，1）值为10
d = ws.cell(row=4, column=1, value=10)
'''



if __name__ == '__main__':
    wb = open_excel(('./javbus.xlsx'))
    ws = open_sheet(wb, 'Demo')
    l=get_value(ws)
    print(l)
