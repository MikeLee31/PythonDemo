import time
from typing import Union

import requests
import os, re, json

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet._write_only import WriteOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet

from Utils import dySpyderUtils
from Utils.ProgressBar import ProgressBar
from Utils import ExcelUtils


def print_users_list_name(users_list):
    for users in users_list:
        print(dySpyderUtils.get_user_name(users))


def dyDownloader(users_list, download_path):
    headers = dySpyderUtils.getdyHeader()

    for user in users_list:
        if user == '':
            print('kong')
            return
        url = user.strip()
        res = requests.get(url, headers=headers, allow_redirects=False)
        sec_uid = re.findall('sec_uid=(.*?)&', res.headers['Location'])[0]

        user_video_url = "https://www.iesdouyin.com/web/api/v2/aweme/post/"
        user_video_params = {
            'sec_uid': sec_uid,
            'count': '21',
            'max_cursor': '0',
            'aid': '1128',
            '_signature': '2Vx9mxAZh0o-K4Wdv7NFKNlcfY',
            'max_cursor': ''
        }

        print('开始获取视频链接------')
        video_urls = []
        max_cursor, video_count, user_name = None, 0, None
        while True:
            # time.sleep(1)
            if max_cursor:
                user_video_params['max_cursor'] = str(max_cursor)
            aweme_count = 0
            while True:
                res = requests.get(user_video_url, headers=headers, params=user_video_params)
                print(res.text)
                js = json.loads(res.text)
                if len(js['aweme_list']) != 0:
                    break
                else:
                    aweme_count = aweme_count + 1
                    time.sleep(5)
                if aweme_count == 3:
                    break

            if aweme_count == 3:
                print("当前用户无作品")
            else:
                if not user_name:
                    user_name = js['aweme_list'][0]['author']['nickname']
                for i in js['aweme_list']:
                    video_count += 1
                    video_urls.append((i['video']['play_addr']['url_list'][0], i['desc']))
                    time.sleep(0.1)
                if js.get('has_more'):
                    max_cursor = js.get('max_cursor')
                else:
                    break

        user_path = download_path + '\\' + user_name
        if not os.path.exists(user_path):
            os.makedirs(user_path)

        num = len(video_urls)
        progressbar = ProgressBar(output='开始下载' + user_name + '的视频', total=num)
        # for i in range(1, 101):
        #     progressbar.move()
        #     time.sleep(0.05)

        cur = 1
        for i in video_urls:
            progressbar.move()
            f_n = i[1]
            url = i[0]
            try:

                f_n = user_path + '\\' + user_name + '-' + f_n + '.mp4'
                print(f_n)
                if os.path.exists(f_n):
                    print('已存在跳过')
                else:
                    time.sleep(1)
                    down_res = requests.get(url=url, headers=headers, verify=False)
                    with open(f_n, "wb") as code:
                        code.write(down_res.content)
            except:
                print('下载失败: ', f_n)
            cur += 1
        print(user_name, '下载完成')


def save_name_to_excel(users_list):
    wb = load_workbook('./users.xlsx')
    ws = wb['users']
    for i in range(0, len(users_list)):
        users_str = users_list[i]
        ind = i + 1
        name = dySpyderUtils.get_user_name(users_str)
        print(name, users_str)
        ws.cell(row=ind,column=1).value=name
        ws.cell(row=ind,column=2).value=users_str

    wb.save('users.xlsx')


if __name__ == '__main__':
    # 填写用户分享链接
    users_list = [
        'https://v.douyin.com/NEugQ4o/'
    ]
    # save_name_to_excel(users_list)
    # 设置下载地址
    download_path = r'E:\Video\dy'
    dyDownloader(users_list, download_path)
    # print_users_list_name(users_list)
