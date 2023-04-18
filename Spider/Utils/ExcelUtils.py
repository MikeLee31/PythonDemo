import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import *


# 打开excel
def open_excel(filepath):
    if os.path.exists(filepath):
        return load_workbook(filepath)
    else:
        Workbook().save(filepath)
        return load_workbook(filepath)


# 打开工作表
def open_sheet(workbook, sheetname):
    if workbook.sheetnames.count(sheetname) > 0:
        ws = workbook[sheetname]
        workbook.remove(ws)
    return workbook.create_sheet(sheetname)


# data为嵌套列表
# li = [
#         ['1', '3'],
#         ['2', '4'],
#     ]
def save_to_excel(datas, excel_name, sheet_name):
    # 打开excel,切换或新建工作表
    filepath = "./" + excel_name + ".xlsx"
    wb = open_excel(filepath)
    print(sheet_name)
    ws = open_sheet(wb, sheet_name)
    index = 1
    for data in datas:
        for j in range(len(data)):
            ws.cell(row=index, column=j + 1).value = data[j]
            print(data[j])
        index += 1
    wb.save(filepath)


if __name__ == '__main__':
    open_excel('./demo.xlsx')
