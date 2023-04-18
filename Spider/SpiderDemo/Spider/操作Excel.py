import openpyxl

wb1 = openpyxl.load_workbook('2019年10月账单.xlsx')

# 获取表
sheet = wb1.worksheets[0]

print(sheet)
# <Worksheet "Sheet2">
print(sheet['A2'].value)
# 遍历数据
for row in sheet.iter_rows():
    for cell in row:
        print(cell.coordinate, cell.value)

# 新建表
wb2 = openpyxl.Workbook()
wb2.create_sheet(index=0, title='小猪')
wb2.create_sheet(index=1, title='佩奇')

# 修改数据以下都可
sheet['A1'].value = '1111111'
sheet.cell(1, 1).value = '1111111'
sheet.cell(1, 1, '111111')
